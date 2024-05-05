from bs4 import BeautifulSoup
import openpyxl, nltk, requests, syllapy, string
from nltk.corpus import stopwords
nltk.download('stopwords')

url = 'https://insights.blackcoffer.com/estimating-the-impact-of-covid-19-on-the-world-of-work-3/'

def remove_words_after_phrase(text, phrase):
    if text != "":
        index = text.find(phrase)
        return text[:index] if index != -1 else text
    
    
def count_cleaned_words(text):

    words = nltk.word_tokenize(text)
    words = [word for word in words if word not in string.punctuation]
    stop_words = set(stopwords.words('english'))
    cleaned_words = [word for word in words if word.lower() not in stop_words]
    word_count = len(cleaned_words)
    cleaned_words_string = ' '.join(cleaned_words)
    
    return word_count, cleaned_words_string

def calculate_positive_score(text, positive_words):
    words = text.split()
    positive_score = sum(1 for word in words if word.lower() in positive_words)
    return positive_score

def calculate_negative_score(text, negative_words):
    words = text.split()
    negative_score = sum(1 for word in words if word.lower() in negative_words)
    return negative_score

def calculate_polarity_score(positive_score, negative_score):
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    return polarity_score

def calculate_subjectivity_score(text, positive_words, negative_words):
    total_words = len(text.split())
    positive_score = calculate_positive_score(text, positive_words)
    negative_score = calculate_negative_score(text, negative_words)
    subjectivity_score = (positive_score + negative_score) / (total_words + 0.000001)
    return subjectivity_score


def calculate_complex_words(text):
    words = text.split()
    total_words = len(words)
    complex_word_count = sum(1 for word in words if syllapy.count(word) > 2)
    if total_words > 0:
        percentage_complex_words = (complex_word_count / total_words) * 100
    else:
        percentage_complex_words = 0
    return percentage_complex_words, complex_word_count

def calculate_average_sentence_length(text):

    sentences = nltk.sent_tokenize(text)
    num_sentences = len(sentences)
    total_words = sum(len(nltk.word_tokenize(sentence)) for sentence in sentences)

    if num_sentences > 0:
        average_sentence_length = total_words / num_sentences
    else:
        average_sentence_length = 0
    
    return average_sentence_length, num_sentences

def calculate_avg_word_per_sentence(word_count,sentence_count):
    return word_count/sentence_count

def calcualte_fog_index(average_length,percentage_complex):
    return 0.4 * (average_length/percentage_complex)

def count_syllables(word):

    exceptions = ["es", "ed"]
    vowels = "aeiou"
    num_syllables = 0
    prev_char_vowel = False
    for char in word:
        if char.lower() in vowels:
            if not prev_char_vowel:
                num_syllables += 1
            prev_char_vowel = True
        else:
            prev_char_vowel = False
    
    if word.endswith(tuple(exceptions)):
        num_syllables -= 1
    
    return max(1, num_syllables)  

def calculate_average_syllable_count(paragraph):
    words = paragraph.split()
    total_syllables = sum(count_syllables(word) for word in words)
    num_words = len(words)
    if num_words > 0:
        average_syllable_count = total_syllables / num_words
    else:
        average_syllable_count = 0
    return average_syllable_count

def count_personal_pronouns(paragraph):

    personal_pronouns = ['i', 'me', 'my', 'mine', 'myself', 'you', 'your', 'yours', 'yourself',
                         'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself',
                         'we', 'us', 'our', 'ours', 'ourselves', 'they', 'them', 'their', 'theirs', 'themselves']
    

    pronoun_count = 0
    
    for word in paragraph.split():
        if word.lower() in personal_pronouns:
            pronoun_count += 1
    
    return pronoun_count

def calculate_average_word_length(paragraph):

    words = paragraph.split()
    total_characters = sum(len(word) for word in words)
    total_words = len(words)
    if total_words > 0:
        average_word_length = total_characters / total_words
    else:
        average_word_length = 0
    
    return average_word_length

def save_data_to_workbook(id,workbook, row_idx, url, positive_score, negative_score, polarity_score, 
                          subjectivity_score, average_length, percentage_complex, fog_index, 
                          average_word_per_sentence, complex_word_count, cleaned_word_count, 
                          average_syllable_count, pronoun_count, average_word_length,file_path):
    sheet = workbook.active
    
    sheet.cell(row=row_idx, column=3, value=positive_score)
    sheet.cell(row=row_idx, column=4, value=negative_score)
    sheet.cell(row=row_idx, column=5, value=polarity_score)
    sheet.cell(row=row_idx, column=6, value=subjectivity_score)
    sheet.cell(row=row_idx, column=7, value=average_length)
    sheet.cell(row=row_idx, column=8, value=percentage_complex)
    sheet.cell(row=row_idx, column=9, value=fog_index)
    sheet.cell(row=row_idx, column=10, value=average_word_per_sentence)
    sheet.cell(row=row_idx, column=11, value=complex_word_count)
    sheet.cell(row=row_idx, column=12, value=cleaned_word_count)
    sheet.cell(row=row_idx, column=13, value=average_syllable_count)
    sheet.cell(row=row_idx, column=14, value=pronoun_count)
    sheet.cell(row=row_idx, column=15, value=average_word_length)
    
    workbook.save(file_path)
    print("")
    print(f"ID: {id}")
    print(f"URL: {url}")
    print("")
    print("1. Positive score:", positive_score)
    print("2. Negative score:", negative_score)
    print("3. Polarity score:", polarity_score)
    print("4. Subjectivity Score:", subjectivity_score)
    print("5. Average sentence length:", average_length)
    print("6. Percentage of complex words:", percentage_complex)
    print("7. FOG index:", fog_index)
    print("8. Average word per sentence:", average_word_per_sentence)
    print("9. Number of complex words:", complex_word_count)
    print("10. Cleaned word count:", cleaned_word_count)
    print("11. Average syllable count per word:", average_syllable_count)
    print("12. Personal pronoun count:", pronoun_count)
    print("13. Average word length:", average_word_length)
    print("")
    print("******************")


#######



def iterate_urls_in_excel(file_path):
    try:

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True), start=2):
            url = row[0]
            celladdress = 'A'+str(row_idx)
            id = sheet[celladdress].value
            if id:
                
                try :
                    html_text = requests.get(url).text
                except:
                    pass
                soup = BeautifulSoup(html_text, 'lxml')
                
                try :
                    title = soup.find('h1', class_='entry-title').text
                except:
                    try:
                        title = soup.find('h1', class_='tdb-title-text').text
                    except:
                        title = ""
                
                if title!= "": 
                    positive_words_file = "positive-words.txt"
                    negative_words_file = "negative-words.txt"

                    with open(positive_words_file, 'r') as f:
                        positive_words = {line.strip().lower() for line in f}

                    with open(negative_words_file, 'r') as f:
                        negative_words = {line.strip().lower() for line in f}
                        
                    try :
                        content = soup.find('div', class_='td-post-content tagdiv-type').text
                    except:
                        try:
                            content = soup.find('div', class_='td_block_wrap tdb_single_content tdi_130 td-pb-border-top td_block_template_1 td-post-content tagdiv-type').div.text
                        except:
                            pass
                        
                    try:
                        phrase = soup.find('pre', 'wp-block-preformatted').text
                        content = remove_words_after_phrase(content, phrase)
                    except:
                        content = content
                        
                    corpus = title + '\n' + content

                    words = corpus.split()
                    num_words = len(words)
                    
                    cleaned_word_count, cleaned_words_string = count_cleaned_words(corpus)
                    positive_score = calculate_positive_score(corpus, positive_words)
                    negative_score = calculate_negative_score(corpus, negative_words)
                    polarity_score = calculate_polarity_score(positive_score, negative_score)
                    subjectivity_score = calculate_subjectivity_score(corpus, positive_words, negative_words)
                    percentage_complex, complex_word_count = calculate_complex_words(corpus)
                    average_length, num_sentences = calculate_average_sentence_length(corpus)
                    average_word_per_sentence = calculate_avg_word_per_sentence(num_words,num_sentences)
                    fog_index = calcualte_fog_index(average_length,percentage_complex)
                    average_syllable_count = calculate_average_syllable_count(corpus)
                    pronoun_count = count_personal_pronouns(corpus)
                    average_word_length = calculate_average_word_length(corpus)
                    
                    save_data_to_workbook(id,workbook, row_idx, url, positive_score, negative_score, 
                                          polarity_score, subjectivity_score, average_length, 
                                          percentage_complex, fog_index, average_word_per_sentence, 
                                          complex_word_count, cleaned_word_count, 
                                          average_syllable_count, pronoun_count, average_word_length,file_path)
                
                
                else:
                    positive_score = 0
                    negative_score = 0
                    polarity_score = 0
                    subjectivity_score = 0
                    average_length = 0
                    percentage_complex = 0
                    fog_index = 0
                    average_word_per_sentence = 0
                    complex_word_count = 0
                    cleaned_word_count = 0
                    average_syllable_count = 0
                    pronoun_count = 0
                    average_word_length = 0 
                    save_data_to_workbook(id,workbook, row_idx, url, positive_score, negative_score, 
                                          polarity_score, subjectivity_score, average_length, 
                                          percentage_complex, fog_index, average_word_per_sentence, 
                                          complex_word_count, cleaned_word_count, 
                                          average_syllable_count, pronoun_count, average_word_length,file_path)
                


    except FileNotFoundError:
        print("File not found.")
    except Exception as e:
        print("An error occurred:", e)

excel_file_path = 'Output Data Structure.xlsx' 
iterate_urls_in_excel(excel_file_path)
